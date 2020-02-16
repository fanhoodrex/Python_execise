import openpyxl 
#写入的代码：
wb = openpyxl.Workbook() 
sheet = wb.active
sheet.title = 'new title'
sheet['A1'] = '漫威宇宙'
rows = [['美国队长','钢铁侠','蜘蛛侠','雷神'],['是','漫威','宇宙', '经典','人物']]
for i in rows:
    sheet.append(i)
print(rows)
wb.save('Marvel.xlsx')

#读取的代码：
wb = openpyxl.load_workbook('Marvel.xlsx')
sheet = wb['new title'] #create object for sheet with name "new title" 
sheetname = wb.sheetnames # list of sheetnames 
print(sheetname)
A1_cell = sheet['A1']
A1_value = A1_cell.value
print(A1_value)
